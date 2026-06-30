from __future__ import annotations

import argparse
import math
import time
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# =========================
# 1. 기본 경로
# =========================
PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT_XLSX = PROJECT_ROOT / "자료" / "버티포트" / "좌표기반_후보지_직선거리_matrix.xlsx"
DEFAULT_CAR_TIME_XLSX = PROJECT_ROOT / "자료" / "버티포트" / "카카오_후보지_자동차_소요시간_matrix.xlsx"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "자료" / "버티포트" 
DEFAULT_OUTPUT_NAME = "JobyS4_UAM_이동소요시간_예측.xlsx"


# =========================
# 2. Joby S4 / 운항 가정
# =========================
JOBY_S4_MAX_SPEED_KMH = 320.0
CRUISE_SPEED_KMH = 300.0
CRUISE_ALTITUDE_M = 600.0
CLIMB_ANGLE_DEG = 8.0
DESCENT_START_RADIUS_KM = 2.0
TAKEOFF_HOVER_MIN = 1.0
LANDING_HOVER_MIN = 1.0


@dataclass(frozen=True)
class FlightAssumptions:
    max_speed_kmh: float = JOBY_S4_MAX_SPEED_KMH
    cruise_speed_kmh: float = CRUISE_SPEED_KMH
    cruise_altitude_m: float = CRUISE_ALTITUDE_M
    climb_angle_deg: float = CLIMB_ANGLE_DEG
    descent_start_radius_km: float = DESCENT_START_RADIUS_KM
    takeoff_hover_min: float = TAKEOFF_HOVER_MIN
    landing_hover_min: float = LANDING_HOVER_MIN

    @property
    def full_climb_ground_distance_km(self) -> float:
        return (self.cruise_altitude_m / 1000.0) / math.tan(math.radians(self.climb_angle_deg))

    @property
    def climb_accel_kmh2(self) -> float:
        """Acceleration implied by 0 -> cruise speed over the full climb distance."""
        return self.cruise_speed_kmh**2 / (2.0 * self.full_climb_ground_distance_km)

    @property
    def nominal_descent_decel_kmh2(self) -> float:
        """Deceleration implied by cruise speed -> 0 over the final descent radius."""
        return self.cruise_speed_kmh**2 / (2.0 * self.descent_start_radius_km)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="좌표기반 직선거리 matrix로 Joby S4 UAM OD별 이동소요시간을 계산합니다."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT_XLSX, help="입력 xlsx 경로")
    parser.add_argument("--car-time-input", type=Path, default=DEFAULT_CAR_TIME_XLSX, help="카카오 자동차 소요시간 xlsx 경로")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="결과 xlsx 저장 폴더")
    parser.add_argument("--output-name", default=DEFAULT_OUTPUT_NAME, help="결과 xlsx 파일명")
    parser.add_argument("--test-limit", type=int, default=None, help="테스트용 OD 행 수 제한")
    parser.add_argument("--cruise-speed-kmh", type=float, default=CRUISE_SPEED_KMH, help="순항/목표 속도")
    return parser.parse_args()


def read_od_distance(input_xlsx: Path) -> tuple[pd.DataFrame, pd.DataFrame | None]:
    if not input_xlsx.exists():
        raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {input_xlsx}")

    try:
        od_df = pd.read_excel(input_xlsx, sheet_name="전체_OD_결과")
        required = {"origin_label", "dest_label", "distance_km"}
        if required.issubset(od_df.columns):
            candidates_df = read_candidates(input_xlsx)
            return normalize_english_od(od_df), candidates_df
    except ValueError:
        pass

    try:
        od_df = pd.read_excel(input_xlsx, sheet_name="OD_컬럼분리")
        required = {"ID_출발지", "ID_도착지", "이동거리_km"}
        if required.issubset(od_df.columns):
            candidates_df = read_candidates(input_xlsx)
            return normalize_korean_od(od_df), candidates_df
    except ValueError:
        pass

    matrix_df = pd.read_excel(input_xlsx, sheet_name="거리_matrix")
    if "origin_label" not in matrix_df.columns:
        raise RuntimeError("거리_matrix 시트에 origin_label 컬럼이 필요합니다.")

    long_df = matrix_df.melt(
        id_vars="origin_label",
        var_name="dest_label",
        value_name="distance_km",
    )
    long_df = long_df.dropna(subset=["distance_km"]).copy()
    long_df = long_df[long_df["origin_label"] != long_df["dest_label"]].reset_index(drop=True)
    long_df["case_no"] = long_df.index + 1
    long_df["route_label"] = long_df["origin_label"].astype(str) + " -> " + long_df["dest_label"].astype(str)
    candidates_df = read_candidates(input_xlsx)
    return long_df, candidates_df


def read_candidates(input_xlsx: Path) -> pd.DataFrame | None:
    try:
        return pd.read_excel(input_xlsx, sheet_name="후보지_목록")
    except ValueError:
        return None


def normalize_english_od(df: pd.DataFrame) -> pd.DataFrame:
    keep_cols = [
        "case_no",
        "total_cases",
        "origin_id",
        "origin_region",
        "origin_label",
        "origin_lon",
        "origin_lat",
        "dest_id",
        "dest_region",
        "dest_label",
        "dest_lon",
        "dest_lat",
        "route_label",
        "distance_km",
        "forward_azimuth_deg",
        "back_azimuth_deg",
        "route_wkt",
        "status",
    ]
    available_cols = [col for col in keep_cols if col in df.columns]
    result = df[available_cols].copy()
    result = result.dropna(subset=["distance_km"]).reset_index(drop=True)
    if "case_no" not in result.columns:
        result["case_no"] = result.index + 1
    if "route_label" not in result.columns:
        result["route_label"] = result["origin_label"].astype(str) + " -> " + result["dest_label"].astype(str)
    return result


def normalize_korean_od(df: pd.DataFrame) -> pd.DataFrame:
    rename_map = {
        "순번": "case_no",
        "전체건수": "total_cases",
        "출발지_ID": "origin_id",
        "출발지": "origin_region",
        "ID_출발지": "origin_label",
        "출발지_경도": "origin_lon",
        "출발지_위도": "origin_lat",
        "도착지_ID": "dest_id",
        "도착지": "dest_region",
        "ID_도착지": "dest_label",
        "도착지_경도": "dest_lon",
        "도착지_위도": "dest_lat",
        "출발지_to_도착지": "route_label",
        "이동거리_km": "distance_km",
        "출발지기준_방위각_deg": "forward_azimuth_deg",
        "도착지기준_역방위각_deg": "back_azimuth_deg",
        "이동경로_WKT": "route_wkt",
        "상태": "status",
    }
    result = df.rename(columns=rename_map).copy()
    result = result.dropna(subset=["distance_km"]).reset_index(drop=True)
    if "case_no" not in result.columns:
        result["case_no"] = result.index + 1
    if "route_label" not in result.columns:
        result["route_label"] = result["origin_label"].astype(str) + " -> " + result["dest_label"].astype(str)
    return result


def kmh_to_km_per_min(speed_kmh: float) -> float:
    return speed_kmh / 60.0


def accelerate_from_rest(distance_km: float, accel_kmh2: float, speed_cap_kmh: float) -> tuple[float, float]:
    """Return (time_min, end_speed_kmh) for constant acceleration from rest."""
    if distance_km <= 0:
        return 0.0, 0.0

    cap_distance_km = speed_cap_kmh**2 / (2.0 * accel_kmh2)
    if distance_km <= cap_distance_km:
        end_speed_kmh = math.sqrt(2.0 * accel_kmh2 * distance_km)
        time_hr = end_speed_kmh / accel_kmh2
        return time_hr * 60.0, end_speed_kmh

    accel_time_hr = speed_cap_kmh / accel_kmh2
    cruise_time_hr = (distance_km - cap_distance_km) / speed_cap_kmh
    return (accel_time_hr + cruise_time_hr) * 60.0, speed_cap_kmh


def short_route_profile(distance_km: float, assumptions: FlightAssumptions) -> dict[str, float | str]:
    if distance_km <= 0:
        return {
            "climb_distance_km": 0.0,
            "cruise_distance_km": 0.0,
            "descent_distance_km": 0.0,
            "climb_time_min": 0.0,
            "cruise_time_min": 0.0,
            "descent_time_min": 0.0,
            "moving_time_min": 0.0,
            "climb_end_altitude_m": 0.0,
            "descent_start_altitude_m": 0.0,
            "climb_end_speed_kmh": 0.0,
            "descent_start_speed_kmh": 0.0,
            "phase_note": "ZERO_DISTANCE",
        }

    accel = assumptions.climb_accel_kmh2
    decel = assumptions.nominal_descent_decel_kmh2
    peak_speed_kmh = math.sqrt((2.0 * distance_km) / ((1.0 / accel) + (1.0 / decel)))
    peak_speed_kmh = min(peak_speed_kmh, assumptions.cruise_speed_kmh)

    climb_distance_km = peak_speed_kmh**2 / (2.0 * accel)
    descent_distance_km = max(distance_km - climb_distance_km, 0.0)
    climb_time_min = (peak_speed_kmh / accel) * 60.0
    descent_time_min = (peak_speed_kmh / decel) * 60.0
    max_altitude_m = min(
        assumptions.cruise_altitude_m,
        climb_distance_km * 1000.0 * math.tan(math.radians(assumptions.climb_angle_deg)),
    )

    return {
        "climb_distance_km": climb_distance_km,
        "cruise_distance_km": 0.0,
        "descent_distance_km": descent_distance_km,
        "climb_time_min": climb_time_min,
        "cruise_time_min": 0.0,
        "descent_time_min": descent_time_min,
        "moving_time_min": climb_time_min + descent_time_min,
        "climb_end_altitude_m": max_altitude_m,
        "descent_start_altitude_m": max_altitude_m,
        "climb_end_speed_kmh": peak_speed_kmh,
        "descent_start_speed_kmh": peak_speed_kmh,
        "phase_note": "SHORT_ROUTE_WITHIN_2KM_RADIUS",
    }


def estimate_flight_time(distance_km: float, assumptions: FlightAssumptions) -> dict[str, float | str]:
    if pd.isna(distance_km):
        return {
            "climb_distance_km": math.nan,
            "cruise_distance_km": math.nan,
            "descent_distance_km": math.nan,
            "climb_time_min": math.nan,
            "cruise_time_min": math.nan,
            "descent_time_min": math.nan,
            "moving_time_min": math.nan,
            "total_time_min": math.nan,
            "climb_end_altitude_m": math.nan,
            "descent_start_altitude_m": math.nan,
            "climb_end_speed_kmh": math.nan,
            "descent_start_speed_kmh": math.nan,
            "phase_note": "MISSING_DISTANCE",
        }

    distance_km = float(distance_km)
    if distance_km <= assumptions.descent_start_radius_km:
        phase = short_route_profile(distance_km, assumptions)
    else:
        available_before_descent_km = max(distance_km - assumptions.descent_start_radius_km, 0.0)
        climb_distance_km = min(assumptions.full_climb_ground_distance_km, available_before_descent_km)
        climb_time_min, climb_end_speed_kmh = accelerate_from_rest(
            climb_distance_km,
            assumptions.climb_accel_kmh2,
            assumptions.cruise_speed_kmh,
        )

        cruise_distance_km = max(available_before_descent_km - climb_distance_km, 0.0)
        cruise_time_min = cruise_distance_km / kmh_to_km_per_min(assumptions.cruise_speed_kmh)

        climb_end_altitude_m = min(
            assumptions.cruise_altitude_m,
            climb_distance_km * 1000.0 * math.tan(math.radians(assumptions.climb_angle_deg)),
        )
        reached_cruise_altitude = math.isclose(
            climb_distance_km,
            assumptions.full_climb_ground_distance_km,
            rel_tol=0.0,
            abs_tol=1e-9,
        )
        descent_start_speed_kmh = assumptions.cruise_speed_kmh if reached_cruise_altitude else climb_end_speed_kmh
        descent_distance_km = assumptions.descent_start_radius_km

        if descent_start_speed_kmh <= 0:
            descent_time_min = 0.0
        else:
            descent_time_min = descent_distance_km / (descent_start_speed_kmh / 2.0 / 60.0)

        phase_note = "FULL_CLIMB_AND_CRUISE" if cruise_distance_km > 0 else "CLIMB_INTERRUPTED_BEFORE_600M"
        if reached_cruise_altitude and cruise_distance_km == 0:
            phase_note = "REACHED_600M_THEN_DESCENT"

        phase = {
            "climb_distance_km": climb_distance_km,
            "cruise_distance_km": cruise_distance_km,
            "descent_distance_km": descent_distance_km,
            "climb_time_min": climb_time_min,
            "cruise_time_min": cruise_time_min,
            "descent_time_min": descent_time_min,
            "moving_time_min": climb_time_min + cruise_time_min + descent_time_min,
            "climb_end_altitude_m": climb_end_altitude_m,
            "descent_start_altitude_m": climb_end_altitude_m,
            "climb_end_speed_kmh": climb_end_speed_kmh,
            "descent_start_speed_kmh": descent_start_speed_kmh,
            "phase_note": phase_note,
        }

    phase["takeoff_hover_min"] = assumptions.takeoff_hover_min
    phase["landing_hover_min"] = assumptions.landing_hover_min
    phase["total_time_min"] = (
        assumptions.takeoff_hover_min + float(phase["moving_time_min"]) + assumptions.landing_hover_min
    )
    return phase


def build_result(od_df: pd.DataFrame, assumptions: FlightAssumptions, test_limit: int | None) -> pd.DataFrame:
    if test_limit is not None:
        od_df = od_df.head(test_limit).copy()

    phase_records = [estimate_flight_time(distance, assumptions) for distance in od_df["distance_km"]]
    phase_df = pd.DataFrame(phase_records)
    result_df = pd.concat([od_df.reset_index(drop=True), phase_df], axis=1)

    if "origin_region" not in result_df.columns:
        result_df["origin_region"] = result_df["origin_label"]
    if "dest_region" not in result_df.columns:
        result_df["dest_region"] = result_df["dest_label"]

    result_df["total_time_sec"] = result_df["total_time_min"] * 60.0
    result_df["moving_time_sec"] = result_df["moving_time_min"] * 60.0
    result_df["average_total_speed_kmh"] = result_df["distance_km"] / (result_df["total_time_min"] / 60.0)
    result_df["average_moving_speed_kmh"] = result_df["distance_km"] / (result_df["moving_time_min"] / 60.0)

    ordered_cols = [
        "case_no",
        "origin_id",
        "origin_region",
        "origin_label",
        "origin_lon",
        "origin_lat",
        "dest_id",
        "dest_region",
        "dest_label",
        "dest_lon",
        "dest_lat",
        "route_label",
        "distance_km",
        "takeoff_hover_min",
        "climb_distance_km",
        "climb_time_min",
        "climb_end_altitude_m",
        "climb_end_speed_kmh",
        "cruise_distance_km",
        "cruise_time_min",
        "descent_distance_km",
        "descent_time_min",
        "descent_start_altitude_m",
        "descent_start_speed_kmh",
        "landing_hover_min",
        "moving_time_min",
        "total_time_min",
        "moving_time_sec",
        "total_time_sec",
        "average_moving_speed_kmh",
        "average_total_speed_kmh",
        "forward_azimuth_deg",
        "back_azimuth_deg",
        "route_wkt",
        "phase_note",
        "status",
    ]
    existing_ordered_cols = [col for col in ordered_cols if col in result_df.columns]
    extra_cols = [col for col in result_df.columns if col not in existing_ordered_cols]
    return result_df[existing_ordered_cols + extra_cols]


def add_car_time_comparison(result_df: pd.DataFrame, car_time_xlsx: Path) -> pd.DataFrame:
    if not car_time_xlsx.exists():
        raise FileNotFoundError(f"자동차 소요시간 파일을 찾을 수 없습니다: {car_time_xlsx}")

    car_df = pd.read_excel(car_time_xlsx)
    required_cols = {"출발지_to_도착지", "소요시간_분"}
    missing_cols = required_cols - set(car_df.columns)
    if missing_cols:
        raise RuntimeError(f"자동차 소요시간 파일에 필요한 컬럼이 없습니다: {sorted(missing_cols)}")

    comparison_df = car_df[["출발지_to_도착지", "소요시간_분"]].copy()
    comparison_df = comparison_df.rename(
        columns={
            "출발지_to_도착지": "route_label",
            "소요시간_분": "car_time_min",
        }
    )
    comparison_df["car_time_min"] = pd.to_numeric(comparison_df["car_time_min"], errors="coerce")

    duplicated_routes = comparison_df["route_label"].duplicated()
    if duplicated_routes.any():
        sample_routes = comparison_df.loc[duplicated_routes, "route_label"].head(5).tolist()
        raise RuntimeError(f"자동차 소요시간 파일에 중복 경로가 있습니다: {sample_routes}")

    car_routes = set(comparison_df["route_label"])
    merged_df = result_df.merge(comparison_df, on="route_label", how="left", validate="one_to_one")
    missing_routes = ~merged_df["route_label"].isin(car_routes) & merged_df["route_label"].notna()
    if missing_routes.any():
        sample_routes = merged_df.loc[missing_routes, "route_label"].head(5).tolist()
        raise RuntimeError(f"자동차 소요시간과 매칭되지 않는 UAM 경로가 있습니다: {sample_routes}")

    merged_df["car_minus_uam_moving_time_min"] = merged_df["car_time_min"] - merged_df["moving_time_min"]
    return merged_df


def make_assumption_table(assumptions: FlightAssumptions) -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("기체", "Joby S4"),
            ("최대 속도 참고값_km/h", assumptions.max_speed_kmh),
            ("계산 적용 순항/목표 속도_km/h", assumptions.cruise_speed_kmh),
            ("순항 고도_m", assumptions.cruise_altitude_m),
            ("상승각_deg", assumptions.climb_angle_deg),
            ("600m 도달 필요 수평거리_km", assumptions.full_climb_ground_distance_km),
            ("하강 시작 반경_km", assumptions.descent_start_radius_km),
            ("이륙 후 Hovering_min", assumptions.takeoff_hover_min),
            ("Hovering & 착륙_min", assumptions.landing_hover_min),
            ("상승 속도 증가", "0에서 목표 속도까지 선형 증가"),
            ("하강 속도 감소", "하강 시작 속도에서 0까지 선형 감소"),
            ("단거리 처리", "출발 시점부터 2km 접근권이면 600m 순항 없이 단거리 가감속 프로파일 적용"),
        ],
        columns=["항목", "값"],
    )


def write_output(
    result_df: pd.DataFrame,
    candidates_df: pd.DataFrame | None,
    assumptions: FlightAssumptions,
    output_path: Path,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    summary_cols = [
        "case_no",
        "origin_label",
        "dest_label",
        "distance_km",
        "total_time_min",
        "moving_time_min",
        "car_time_min",
        "car_minus_uam_moving_time_min",
        "climb_end_altitude_m",
        "phase_note",
    ]
    summary_cols = [col for col in summary_cols if col in result_df.columns]
    summary_df = result_df[summary_cols].copy()

    total_time_matrix = result_df.pivot(
        index="origin_label",
        columns="dest_label",
        values="total_time_min",
    )
    moving_time_matrix = result_df.pivot(
        index="origin_label",
        columns="dest_label",
        values="moving_time_min",
    )
    altitude_matrix = result_df.pivot(
        index="origin_label",
        columns="dest_label",
        values="climb_end_altitude_m",
    )

    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            make_assumption_table(assumptions).to_excel(writer, sheet_name="입력가정", index=False)
            summary_df.to_excel(writer, sheet_name="요약", index=False)
            result_df.to_excel(writer, sheet_name="전체_OD_시간", index=False)
            total_time_matrix.to_excel(writer, sheet_name="총소요시간_matrix")
            moving_time_matrix.to_excel(writer, sheet_name="이동시간_matrix")
            altitude_matrix.to_excel(writer, sheet_name="최대고도_matrix")
            if candidates_df is not None:
                candidates_df.to_excel(writer, sheet_name="후보지_목록", index=False)
            format_workbook(writer.book)
        return output_path
    except PermissionError:
        alt_path = output_path.with_name(f"{output_path.stem}_{int(time.time())}{output_path.suffix}")
        with pd.ExcelWriter(alt_path, engine="openpyxl") as writer:
            make_assumption_table(assumptions).to_excel(writer, sheet_name="입력가정", index=False)
            summary_df.to_excel(writer, sheet_name="요약", index=False)
            result_df.to_excel(writer, sheet_name="전체_OD_시간", index=False)
            total_time_matrix.to_excel(writer, sheet_name="총소요시간_matrix")
            moving_time_matrix.to_excel(writer, sheet_name="이동시간_matrix")
            altitude_matrix.to_excel(writer, sheet_name="최대고도_matrix")
            if candidates_df is not None:
                candidates_df.to_excel(writer, sheet_name="후보지_목록", index=False)
            format_workbook(writer.book)
        return alt_path


def format_workbook(workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = 0
            for cell in column_cells:
                if cell.value is None:
                    continue
                max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 32)

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.000"


def main() -> None:
    args = parse_args()
    assumptions = FlightAssumptions(cruise_speed_kmh=args.cruise_speed_kmh)
    if assumptions.cruise_speed_kmh > assumptions.max_speed_kmh:
        raise ValueError("계산 적용 순항/목표 속도는 Joby S4 최대 속도보다 클 수 없습니다.")

    od_df, candidates_df = read_od_distance(args.input)
    result_df = build_result(od_df, assumptions, args.test_limit)
    result_df = add_car_time_comparison(result_df, args.car_time_input)

    output_path = args.output_dir / args.output_name
    saved_path = write_output(result_df, candidates_df, assumptions, output_path)

    print(f"입력 OD 수: {len(result_df):,}")
    print(f"600m 도달 필요 수평거리: {assumptions.full_climb_ground_distance_km:.3f} km")
    print(f"결과 저장 완료: {saved_path}")


if __name__ == "__main__":
    main()