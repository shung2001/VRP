from __future__ import annotations

import argparse
import time
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_ROOT / "자료" / "버티포트"

DEFAULT_INPUT_XLSX = DATA_DIR / "S4_UAM_이동소요시간_예측.xlsx"
FALLBACK_INPUT_XLSX = DATA_DIR / "JobyS4_UAM_이동소요시간_예측.xlsx"
DEFAULT_OUTPUT_XLSX = DATA_DIR / "totla_car_minus_uam_moving_time.xlsx"

GROUP_COL = "origin_label"
VALUE_COL = "car_minus_uam_moving_time_min"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "UAM 이동소요시간 예측 결과에서 origin_label별 "
            "car_minus_uam_moving_time_min 합계를 계산합니다."
        )
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT_XLSX, help="입력 xlsx 경로")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_XLSX, help="출력 xlsx 경로")
    parser.add_argument("--sheet", default=None, help="입력 시트명. 생략하면 필요한 컬럼이 있는 시트를 자동 탐색")
    return parser.parse_args()


def resolve_input_path(input_path: Path) -> Path:
    if input_path.exists():
        return input_path
    if input_path == DEFAULT_INPUT_XLSX and FALLBACK_INPUT_XLSX.exists():
        print(f"기본 입력 파일이 없어 대체 파일을 사용합니다: {FALLBACK_INPUT_XLSX}")
        return FALLBACK_INPUT_XLSX
    raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {input_path}")


def find_sheet_name(input_path: Path, requested_sheet: str | None) -> str:
    excel = pd.ExcelFile(input_path)

    if requested_sheet is not None:
        if requested_sheet not in excel.sheet_names:
            raise RuntimeError(f"'{requested_sheet}' 시트가 없습니다. 사용 가능 시트: {excel.sheet_names}")
        return requested_sheet

    for sheet_name in ["요약", "전체_OD_시간", *excel.sheet_names]:
        if sheet_name not in excel.sheet_names:
            continue
        headers = pd.read_excel(input_path, sheet_name=sheet_name, nrows=0).columns
        if GROUP_COL in headers and VALUE_COL in headers:
            return sheet_name

    raise RuntimeError(
        f"'{GROUP_COL}'와 '{VALUE_COL}' 컬럼이 함께 있는 시트를 찾지 못했습니다. "
        f"사용 가능 시트: {excel.sheet_names}"
    )


def aggregate_by_origin(input_path: Path, sheet_name: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    df = pd.read_excel(input_path, sheet_name=sheet_name)

    missing_cols = [col for col in [GROUP_COL, VALUE_COL] if col not in df.columns]
    if missing_cols:
        raise RuntimeError(f"입력 시트에 필요한 컬럼이 없습니다: {missing_cols}")

    work_df = df[[GROUP_COL, VALUE_COL]].copy()
    work_df[VALUE_COL] = pd.to_numeric(work_df[VALUE_COL], errors="coerce")
    work_df = work_df.dropna(subset=[GROUP_COL])
    invalid_count = int(work_df[VALUE_COL].isna().sum())

    result_df = (
        work_df.groupby(GROUP_COL, as_index=False)
        .agg(
            total_car_minus_uam_moving_time_min=(VALUE_COL, lambda values: values.sum(min_count=1)),
            input_od_count=(VALUE_COL, "size"),
            valid_value_count=(VALUE_COL, "count"),
            average_car_minus_uam_moving_time_min=(VALUE_COL, "mean"),
        )
        .sort_values("total_car_minus_uam_moving_time_min", ascending=False, na_position="last")
        .reset_index(drop=True)
    )

    result_df["missing_value_count"] = result_df["input_od_count"] - result_df["valid_value_count"]
    result_df.insert(0, "rank", result_df.index + 1)

    metadata_df = pd.DataFrame(
        [
            ("입력 파일", str(input_path)),
            ("입력 시트", sheet_name),
            ("그룹 컬럼", GROUP_COL),
            ("합계 컬럼", VALUE_COL),
            ("입력 행 수", len(df)),
            ("origin_label 존재 행 수", len(work_df)),
            ("집계 유효값 행 수", int(result_df["valid_value_count"].sum())),
            ("숫자 변환 실패/결측 행 수", invalid_count),
            ("origin_label 그룹 수", len(result_df)),
        ],
        columns=["항목", "값"],
    )

    return result_df, metadata_df


def save_workbook(result_df: pd.DataFrame, metadata_df: pd.DataFrame, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        saved_path = output_path
        with pd.ExcelWriter(saved_path, engine="openpyxl") as writer:
            result_df.to_excel(writer, sheet_name="origin별_합계", index=False)
            metadata_df.to_excel(writer, sheet_name="집계정보", index=False)
            format_workbook(writer.book)
        return saved_path
    except PermissionError:
        saved_path = output_path.with_name(f"{output_path.stem}_{int(time.time())}{output_path.suffix}")
        with pd.ExcelWriter(saved_path, engine="openpyxl") as writer:
            result_df.to_excel(writer, sheet_name="origin별_합계", index=False)
            metadata_df.to_excel(writer, sheet_name="집계정보", index=False)
            format_workbook(writer.book)
        return saved_path


def format_workbook(workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = 0
            for cell in column_cells:
                if cell.value is None:
                    continue
                max_length = max(max_length, len(str(cell.value)))
                if isinstance(cell.value, float):
                    cell.number_format = "0.000"
                elif isinstance(cell.value, int):
                    cell.number_format = "#,##0"

            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 42)


def main() -> None:
    args = parse_args()
    input_path = resolve_input_path(args.input)
    sheet_name = find_sheet_name(input_path, args.sheet)
    result_df, metadata_df = aggregate_by_origin(input_path, sheet_name)
    saved_path = save_workbook(result_df, metadata_df, args.output)

    print(f"입력 파일: {input_path}")
    print(f"입력 시트: {sheet_name}")
    print(f"origin_label 그룹 수: {len(result_df):,}")
    print(f"결과 저장 완료: {saved_path}")


if __name__ == "__main__":
    main()
