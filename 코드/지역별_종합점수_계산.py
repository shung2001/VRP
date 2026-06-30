from __future__ import annotations

import argparse
import re
import time
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pyproj import Transformer


PROJECT_ROOT = Path(__file__).resolve().parents[1]

DATA_DIR_NAME = "자료"
CODE_DIR_NAME = "코드"
VERTIPORT_DIR_NAME = "버티포트"
RESULT_DIR_NAME = "지역별_결과"

DEFAULT_TIME_XLSX = PROJECT_ROOT / DATA_DIR_NAME / VERTIPORT_DIR_NAME / "total_car_minus_uam_moving_time.xlsx"
DEFAULT_COORD_XLSX = (
    PROJECT_ROOT
    / DATA_DIR_NAME
    / VERTIPORT_DIR_NAME
    / "각_지역별_버티포트_후보군(중심점)_주소추가_최종.xlsx"
)
DEFAULT_VALUE_CSV = PROJECT_ROOT / DATA_DIR_NAME / RESULT_DIR_NAME / "SI_DO_value_개수_시단위.csv"
DEFAULT_OUTPUT_CSV = PROJECT_ROOT / DATA_DIR_NAME / RESULT_DIR_NAME / "지역별_종합점수.csv"
DEFAULT_OUTPUT_XLSX = PROJECT_ROOT / DATA_DIR_NAME / RESULT_DIR_NAME / "지역별_종합점수.xlsx"

TIME_REGION_COL = "origin_label"
TIME_VALUE_COL = "total_car_minus_uam_moving_time_min"
TIME_ID_COL = "origin_id"

VALUE_ID_COL = "CITY_CD"
VALUE_REGION_COL = "CITY_NM"
VALUE_COLUMNS = ["Company", "MICE", "Korea_Tourist", "Shopping_Mall"]
VALUE_COLUMN_ALIASES = {
    "company": "Company",
    "mice": "MICE",
    "korea_tourist": "Korea_Tourist",
    "shopping_mall": "Shopping_Mall",
    "Shopping_MAll": "Shopping_Mall",
    "Shopping_Mall": "Shopping_Mall",
}

COORD_REGION_COL = "지역"
COORD_ADDRESS_COL = "주소"
COORD_LON_COL = "경도"
COORD_LAT_COL = "위도"
COORD_X_COL = "MEAN_X"
COORD_Y_COL = "MEAN_Y"

# Edit these values when the score weighting needs to change.
# The script normalizes weights automatically, so they do not need to sum to 1.
WEIGHTS = {
    TIME_VALUE_COL: 0.40,
    "Company": 0.20,
    "MICE": 0.15,
    "Korea_Tourist": 0.15,
    "Shopping_Mall": 0.10,
}

# Add manual region-name fixes here when the two input files use different names.
# Example: {"기흥구": "용인시"}
REGION_NAME_REPLACEMENTS: dict[str, str] = {}

# left keeps only registered vertiport candidates from the moving-time file.
MERGE_HOW = "left"
SCORE_SCALE = 100.0

SIDO_CODE_BY_ADDRESS_PREFIX = {
    "서울": "11",
    "서울특별시": "11",
    "인천": "23",
    "인천광역시": "23",
    "경기": "31",
    "경기도": "31",
    "충남": "34",
    "충청남도": "34",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Calculate weighted regional composite scores using registered vertiport "
            "candidate IDs and write coordinates to a separate Excel sheet."
        )
    )
    parser.add_argument("--time-xlsx", type=Path, default=DEFAULT_TIME_XLSX)
    parser.add_argument("--coord-xlsx", type=Path, default=DEFAULT_COORD_XLSX)
    parser.add_argument("--value-csv", type=Path, default=DEFAULT_VALUE_CSV)
    parser.add_argument("--time-sheet", default=None, help="Excel sheet name. Auto-detected when omitted.")
    parser.add_argument("--coord-sheet", default=None, help="Coordinate sheet name. Auto-detected when omitted.")
    parser.add_argument("--output-csv", type=Path, default=DEFAULT_OUTPUT_CSV)
    parser.add_argument("--output-xlsx", type=Path, default=DEFAULT_OUTPUT_XLSX)
    parser.add_argument("--no-xlsx", action="store_true", help="Only write CSV output.")
    parser.add_argument("--merge-how", choices=["left", "inner", "outer", "right"], default=MERGE_HOW)
    parser.add_argument(
        "--weights",
        default=None,
        help=(
            "Optional overrides, e.g. "
            "'time=0.5,Company=0.2,MICE=0.1,Korea_Tourist=0.1,Shopping_Mall=0.1'"
        ),
    )
    return parser.parse_args()


def resolve_path(path: Path, fallback_name: str) -> Path:
    if path.exists():
        return path

    matches = [p for p in PROJECT_ROOT.rglob(fallback_name) if not p.name.startswith("~$")]
    if len(matches) == 1:
        return matches[0]
    if matches:
        return sorted(matches, key=path_priority)[0]

    raise FileNotFoundError(f"Input file not found: {path}")


def path_priority(path: Path) -> tuple[int, int, str]:
    parts = path.relative_to(PROJECT_ROOT).parts
    score = 0

    if parts and parts[0] == DATA_DIR_NAME:
        score -= 100
    if parts and parts[0] == CODE_DIR_NAME:
        score += 100
    if RESULT_DIR_NAME in parts:
        score -= 20

    return score, len(parts), str(path)


def read_csv_with_fallback(path: Path) -> pd.DataFrame:
    errors: list[str] = []
    for encoding in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            return pd.read_csv(path, encoding=encoding)
        except UnicodeDecodeError as exc:
            errors.append(f"{encoding}: {exc}")
    raise RuntimeError(f"Could not read CSV with known encodings: {path}\n" + "\n".join(errors))


def find_time_sheet(path: Path, requested_sheet: str | None) -> str:
    excel = pd.ExcelFile(path)

    if requested_sheet is not None:
        if requested_sheet not in excel.sheet_names:
            raise RuntimeError(f"Sheet not found: {requested_sheet}. Available sheets: {excel.sheet_names}")
        return requested_sheet

    for sheet_name in excel.sheet_names:
        headers = pd.read_excel(path, sheet_name=sheet_name, nrows=0).columns
        if TIME_REGION_COL in headers and TIME_VALUE_COL in headers:
            return sheet_name

    raise RuntimeError(
        f"No sheet contains required columns: {TIME_REGION_COL}, {TIME_VALUE_COL}. "
        f"Available sheets: {excel.sheet_names}"
    )


def find_coordinate_sheet(path: Path, requested_sheet: str | None) -> str:
    excel = pd.ExcelFile(path)
    required_cols = [COORD_REGION_COL, COORD_LON_COL, COORD_LAT_COL]

    if requested_sheet is not None:
        if requested_sheet not in excel.sheet_names:
            raise RuntimeError(f"Sheet not found: {requested_sheet}. Available sheets: {excel.sheet_names}")
        headers = pd.read_excel(path, sheet_name=requested_sheet, nrows=0).columns
        missing = [col for col in required_cols if col not in headers]
        if missing:
            raise RuntimeError(f"Missing coordinate columns in {requested_sheet}: {missing}")
        return requested_sheet

    for sheet_name in excel.sheet_names:
        headers = pd.read_excel(path, sheet_name=sheet_name, nrows=0).columns
        if all(col in headers for col in required_cols):
            return sheet_name

    raise RuntimeError(
        f"No sheet contains required coordinate columns: {required_cols}. "
        f"Available sheets: {excel.sheet_names}"
    )


def normalize_region_name(value: object) -> str | None:
    if pd.isna(value):
        return None

    name = str(value).strip()
    if not name:
        return None

    name = re.sub(r"^\d+_", "", name)
    name = re.sub(r"_\d+$", "", name)
    return REGION_NAME_REPLACEMENTS.get(name, name)


def extract_origin_id(value: object) -> str | None:
    if pd.isna(value):
        return None

    match = re.match(r"^(\d+)_", str(value).strip())
    if match is None:
        return None
    return match.group(1)


def format_origin_label(origin_id: int, region: object) -> str:
    return f"{origin_id:02d}_{str(region).strip()}"


def infer_sido_code(address: object) -> str | None:
    if pd.isna(address):
        return None

    text = str(address).strip()
    for prefix, code in sorted(SIDO_CODE_BY_ADDRESS_PREFIX.items(), key=lambda item: len(item[0]), reverse=True):
        if text.startswith(prefix):
            return code
    return None


def extract_value_sido_code(value: object) -> str | None:
    if pd.isna(value):
        return None

    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    if len(digits) < 2:
        return None
    return digits[:2]


def load_coordinate_table(path: Path, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet_name)
    required_cols = [COORD_REGION_COL, COORD_LON_COL, COORD_LAT_COL]
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        raise RuntimeError(f"Missing coordinate columns: {missing}")

    work = df.copy()
    work.insert(0, TIME_ID_COL, range(1, len(work) + 1))
    work["coord_region_raw"] = work[COORD_REGION_COL].astype(str).str.strip()
    work[TIME_REGION_COL] = [
        format_origin_label(origin_id, region)
        for origin_id, region in zip(work[TIME_ID_COL], work["coord_region_raw"], strict=False)
    ]
    work["region_name"] = work["coord_region_raw"].map(normalize_region_name)
    work["lon_4326"] = pd.to_numeric(work[COORD_LON_COL], errors="coerce")
    work["lat_4326"] = pd.to_numeric(work[COORD_LAT_COL], errors="coerce")

    transformer = Transformer.from_crs("EPSG:4326", "EPSG:5186", always_xy=True)
    valid_xy = work["lon_4326"].notna() & work["lat_4326"].notna()
    work["x_5186"] = pd.NA
    work["y_5186"] = pd.NA
    if valid_xy.any():
        xs, ys = transformer.transform(
            work.loc[valid_xy, "lon_4326"].to_numpy(),
            work.loc[valid_xy, "lat_4326"].to_numpy(),
        )
        work.loc[valid_xy, "x_5186"] = xs
        work.loc[valid_xy, "y_5186"] = ys

    if COORD_ADDRESS_COL in work.columns:
        work["address"] = work[COORD_ADDRESS_COL]
        work["sido_code"] = work[COORD_ADDRESS_COL].map(infer_sido_code)
    else:
        work["address"] = pd.NA
        work["sido_code"] = pd.NA

    if COORD_X_COL in work.columns:
        work["source_x_5186"] = pd.to_numeric(work[COORD_X_COL], errors="coerce")
    else:
        work["source_x_5186"] = pd.NA

    if COORD_Y_COL in work.columns:
        work["source_y_5186"] = pd.to_numeric(work[COORD_Y_COL], errors="coerce")
    else:
        work["source_y_5186"] = pd.NA

    output_cols = [
        TIME_ID_COL,
        TIME_REGION_COL,
        "region_name",
        "coord_region_raw",
        "address",
        "sido_code",
        "lon_4326",
        "lat_4326",
        "x_5186",
        "y_5186",
        "source_x_5186",
        "source_y_5186",
    ]
    return work[output_cols].reset_index(drop=True)


def load_time_scores(
    path: Path,
    sheet_name: str,
    coordinate_df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    df = pd.read_excel(path, sheet_name=sheet_name)
    missing = [col for col in [TIME_REGION_COL, TIME_VALUE_COL] if col not in df.columns]
    if missing:
        raise RuntimeError(f"Missing columns in time Excel: {missing}")

    work = df[[TIME_REGION_COL, TIME_VALUE_COL]].copy()
    work.insert(0, "time_row_id", range(1, len(work) + 1))
    work[TIME_ID_COL] = work[TIME_REGION_COL].map(extract_origin_id)
    work["region_name"] = work[TIME_REGION_COL].map(normalize_region_name)
    work[TIME_VALUE_COL] = pd.to_numeric(work[TIME_VALUE_COL], errors="coerce")

    coord_key = coordinate_df[[TIME_REGION_COL, "sido_code"]].rename(columns={"sido_code": "coord_sido_code"})
    work = work.merge(coord_key, on=TIME_REGION_COL, how="left")

    invalid_rows = work[work["region_name"].isna() | work[TIME_VALUE_COL].isna()].copy()
    work = work.dropna(subset=["region_name", TIME_VALUE_COL])
    work = work.sort_values(["region_name", TIME_ID_COL, "time_row_id"], na_position="last").reset_index(drop=True)
    return work, invalid_rows


def load_value_scores(path: Path) -> pd.DataFrame:
    df = read_csv_with_fallback(path)
    df = standardize_value_columns(df)
    required_cols = [VALUE_ID_COL, VALUE_REGION_COL, *VALUE_COLUMNS]
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        raise RuntimeError(f"Missing columns in value CSV: {missing}")

    work = df[required_cols].copy()
    work.insert(0, "value_row_id", range(1, len(work) + 1))
    work["region_name"] = work[VALUE_REGION_COL].map(normalize_region_name)
    work["value_sido_code"] = work[VALUE_ID_COL].map(extract_value_sido_code)
    for col in VALUE_COLUMNS:
        work[col] = pd.to_numeric(work[col], errors="coerce")

    work = work.dropna(subset=["region_name"])
    return work.sort_values(["region_name", VALUE_ID_COL, "value_row_id"]).reset_index(drop=True)


def standardize_value_columns(df: pd.DataFrame) -> pd.DataFrame:
    rename_map = {}
    for col in df.columns:
        stripped = str(col).strip()
        canonical = VALUE_COLUMN_ALIASES.get(stripped)
        if canonical is None:
            canonical = VALUE_COLUMN_ALIASES.get(stripped.lower())
        if canonical is not None:
            rename_map[col] = canonical

    return df.rename(columns=rename_map)


def parse_weights(overrides: str | None) -> dict[str, float]:
    weights = dict(WEIGHTS)
    aliases = {
        "time": TIME_VALUE_COL,
        "uam_time": TIME_VALUE_COL,
        "moving_time": TIME_VALUE_COL,
        "time_saving": TIME_VALUE_COL,
        "Shopping_MAll": "Shopping_Mall",
    }

    if overrides:
        for item in overrides.split(","):
            if not item.strip():
                continue
            if "=" not in item:
                raise RuntimeError(f"Invalid weight override: {item}")
            key, value = item.split("=", 1)
            key = aliases.get(key.strip(), key.strip())
            if key not in weights:
                raise RuntimeError(f"Unknown weight column: {key}. Available columns: {list(weights)}")
            weights[key] = float(value)

    if any(value < 0 for value in weights.values()):
        raise RuntimeError(f"Weights must be non-negative: {weights}")

    total = sum(weights.values())
    if total <= 0:
        raise RuntimeError(f"At least one weight must be greater than zero: {weights}")

    return {key: value / total for key, value in weights.items()}


def minmax_score(series: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce").fillna(0)
    min_value = numeric.min()
    max_value = numeric.max()

    if max_value == min_value:
        fill_value = SCORE_SCALE if max_value > 0 else 0.0
        return pd.Series(fill_value, index=series.index)

    return (numeric - min_value) / (max_value - min_value) * SCORE_SCALE


def calculate_composite_score(
    time_df: pd.DataFrame,
    value_df: pd.DataFrame,
    weights: dict[str, float],
    merge_how: str,
) -> pd.DataFrame:
    merged = pd.merge(
        time_df,
        value_df,
        left_on=["region_name", "coord_sido_code"],
        right_on=["region_name", "value_sido_code"],
        how=merge_how,
    )

    merged["time_present"] = merged[TIME_REGION_COL].notna() if TIME_REGION_COL in merged.columns else False
    merged["value_present"] = merged[VALUE_REGION_COL].notna() if VALUE_REGION_COL in merged.columns else False

    score_columns = [TIME_VALUE_COL, *VALUE_COLUMNS]
    for col in score_columns:
        merged[col] = pd.to_numeric(merged[col], errors="coerce").fillna(0)

    total_score = pd.Series(0.0, index=merged.index)
    for col in score_columns:
        normalized_col = f"{col}_score_0_100"
        weighted_col = f"{col}_weighted_score"
        merged[normalized_col] = minmax_score(merged[col])
        merged[weighted_col] = merged[normalized_col] * weights[col]
        total_score += merged[weighted_col]

    merged["total_score"] = total_score
    merged = merged.sort_values(["total_score", TIME_REGION_COL], ascending=[False, True]).reset_index(drop=True)
    merged.insert(0, "rank", range(1, len(merged) + 1))

    preferred_cols = [
        "rank",
        "region_name",
        "total_score",
        "time_row_id",
        TIME_ID_COL,
        TIME_REGION_COL,
        "coord_sido_code",
        "value_row_id",
        VALUE_ID_COL,
        VALUE_REGION_COL,
        "value_sido_code",
        TIME_VALUE_COL,
        *VALUE_COLUMNS,
        *[f"{col}_score_0_100" for col in score_columns],
        *[f"{col}_weighted_score" for col in score_columns],
        "time_present",
        "value_present",
    ]
    existing_cols = [col for col in preferred_cols if col in merged.columns]
    other_cols = [col for col in merged.columns if col not in existing_cols]
    return merged[existing_cols + other_cols]


def build_weight_table(weights: dict[str, float]) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "column": list(weights.keys()),
            "normalized_weight": list(weights.values()),
        }
    )


def save_outputs(
    result_df: pd.DataFrame,
    time_df: pd.DataFrame,
    value_df: pd.DataFrame,
    coordinate_df: pd.DataFrame,
    weights_df: pd.DataFrame,
    invalid_time_rows: pd.DataFrame,
    output_csv: Path,
    output_xlsx: Path,
    write_xlsx: bool,
) -> tuple[Path, Path | None]:
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)

    csv_path = output_csv
    try:
        result_df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    except PermissionError:
        csv_path = output_csv.with_name(f"{output_csv.stem}_{int(time.time())}{output_csv.suffix}")
        result_df.to_csv(csv_path, index=False, encoding="utf-8-sig")

    if not write_xlsx:
        return csv_path, None

    if {"time_present", "value_present"}.issubset(result_df.columns):
        unmatched_df = result_df[~(result_df["time_present"] & result_df["value_present"])].copy()
    else:
        unmatched_df = pd.DataFrame()

    xlsx_path = output_xlsx
    try:
        write_workbook(result_df, coordinate_df, time_df, value_df, weights_df, unmatched_df, invalid_time_rows, xlsx_path)
    except PermissionError:
        xlsx_path = output_xlsx.with_name(f"{output_xlsx.stem}_{int(time.time())}{output_xlsx.suffix}")
        write_workbook(result_df, coordinate_df, time_df, value_df, weights_df, unmatched_df, invalid_time_rows, xlsx_path)

    return csv_path, xlsx_path


def write_workbook(
    result_df: pd.DataFrame,
    coordinate_df: pd.DataFrame,
    time_df: pd.DataFrame,
    value_df: pd.DataFrame,
    weights_df: pd.DataFrame,
    unmatched_df: pd.DataFrame,
    invalid_time_rows: pd.DataFrame,
    output_path: Path,
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        result_df.to_excel(writer, sheet_name="regional_score", index=False)
        coordinate_df.to_excel(writer, sheet_name="coordinates", index=False)
        time_df.to_excel(writer, sheet_name="time_input_rows", index=False)
        value_df.to_excel(writer, sheet_name="value_input_rows", index=False)
        weights_df.to_excel(writer, sheet_name="weights", index=False)
        unmatched_df.to_excel(writer, sheet_name="not_matched", index=False)
        invalid_time_rows.to_excel(writer, sheet_name="invalid_time_rows", index=False)
        format_workbook(writer.book)


def format_workbook(workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = 0
            for cell in column_cells:
                if cell.value is None:
                    continue
                max_length = max(max_length, len(str(cell.value)))
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000"
                elif isinstance(cell.value, int):
                    cell.number_format = "#,##0"

            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 48)


def main() -> None:
    args = parse_args()

    time_xlsx = resolve_path(args.time_xlsx, "total_car_minus_uam_moving_time.xlsx")
    coord_xlsx = resolve_path(args.coord_xlsx, "각_지역별_버티포트_후보군(중심점)_주소추가_최종.xlsx")
    value_csv = resolve_path(args.value_csv, "SI_DO_value_개수_시단위.csv")
    time_sheet_name = find_time_sheet(time_xlsx, args.time_sheet)
    coord_sheet_name = find_coordinate_sheet(coord_xlsx, args.coord_sheet)
    weights = parse_weights(args.weights)

    coordinate_df = load_coordinate_table(coord_xlsx, coord_sheet_name)
    time_df, invalid_time_rows = load_time_scores(time_xlsx, time_sheet_name, coordinate_df)
    value_df = load_value_scores(value_csv)
    result_df = calculate_composite_score(time_df, value_df, weights, args.merge_how)
    weights_df = build_weight_table(weights)
    csv_path, xlsx_path = save_outputs(
        result_df=result_df,
        time_df=time_df,
        value_df=value_df,
        coordinate_df=coordinate_df,
        weights_df=weights_df,
        invalid_time_rows=invalid_time_rows,
        output_csv=args.output_csv,
        output_xlsx=args.output_xlsx,
        write_xlsx=not args.no_xlsx,
    )

    complete_count = int((result_df["time_present"] & result_df["value_present"]).sum())
    separated_count = len(result_df) - complete_count

    print(f"time_xlsx: {time_xlsx}")
    print(f"time_sheet: {time_sheet_name}")
    print(f"coord_xlsx: {coord_xlsx}")
    print(f"coord_sheet: {coord_sheet_name}")
    print(f"value_csv: {value_csv}")
    print(f"rows: {len(result_df):,} / complete_rows: {complete_count:,} / separated_rows: {separated_count:,}")
    print(f"csv_output: {csv_path}")
    if xlsx_path is not None:
        print(f"xlsx_output: {xlsx_path}")


if __name__ == "__main__":
    main()
